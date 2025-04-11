from django.shortcuts import render
from django.views.generic import TemplateView
from django.db.models import Count, Q
from django.utils import timezone
from users.models import Booking
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from io import BytesIO

class BookingsOverviewReport(TemplateView):
    template_name = 'reports/bookings_overview.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Default filter: Current month
        month = self.request.GET.get('month', timezone.now().month)
        year = self.request.GET.get('year', timezone.now().year)
        service_type = self.request.GET.get('service_type')

        # Base queryset
        bookings = Booking.objects.filter(
            created_at__year=year,
            created_at__month=month
        )

        # Apply service type filter if selected
        if service_type:
            bookings = bookings.filter(service=service_type)

        # Aggregate counts
        total = bookings.count()
        pending = bookings.filter(status='bidding').count() + bookings.filter(status='confirmed').count()
        completed = bookings.filter(status='completed').count()
        cancelled = bookings.filter(status='cancelled').count()
        

        # Pass data to template
        context.update({
            'total': total,
            'pending': pending,
            'completed': completed,
            'cancelled': cancelled,
            'year': year,
            'month': month,
            'service_type': service_type,
            # 'service_types': Booking.SERVICE,  # For filter dropdown
        })
        return context
    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        
        if export_format in ['pdf', 'excel']:
            return self.export_report(export_format)
        return super().get(request, *args, **kwargs)

    def export_report(self, export_format):
        # Get the data you need for export
        context = self.get_context_data()
        data = [
            ['Metric', 'Count'],
            ['Total Bookings', context['total']],
            ['Pending', context['pending']],
            ['Completed', context['completed']],
            ['Cancelled', context['cancelled']]
        ]
        
        if export_format == 'pdf':
            return self.export_to_pdf(data)
        elif export_format == 'excel':
            return self.export_to_excel(data)

    def export_to_pdf(self, data):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        styles = getSampleStyleSheet()
        title = Paragraph("Bookings Overview Report", styles['Title'])
        elements.append(title)
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="bookings_report.pdf"'
        return response

    def export_to_excel(self, data):
        try:
            df = pd.DataFrame(data[1:], columns=data[0])
            
            with BytesIO() as buffer:
                # Using context manager for proper resource handling
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df.to_excel(
                        writer,
                        index=False,
                        sheet_name='Bookings Report',
                        freeze_panes=(1, 0)  # Freeze header row
                    )
                    worksheet = writer.sheets['Bookings Report']
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                    
                    # Format header
                    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    header_font = Font(color='FFFFFF', bold=True)
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font

                buffer.seek(0)
                response = HttpResponse(
                    buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename="bookings_report.xlsx"'
                return response
                
        except Exception as e:
            return HttpResponse(f"Error generating Excel file: {str(e)}", status=500)